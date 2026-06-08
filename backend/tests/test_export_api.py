from io import BytesIO
from pathlib import Path
import sys
import zipfile

from openpyxl import load_workbook


TESTS_DIR = Path(__file__).resolve().parent
BACKEND_DIR = TESTS_DIR.parent
ROOT_DIR = BACKEND_DIR.parent
PROTOTYPE_SRC_DIR = ROOT_DIR / "prototype" / "src"
if str(TESTS_DIR) not in sys.path:
    sys.path.insert(0, str(TESTS_DIR))
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))
if str(PROTOTYPE_SRC_DIR) not in sys.path:
    sys.path.insert(0, str(PROTOTYPE_SRC_DIR))

from test_api import client, setup_function as base_setup_function


STATUS_TODO = "\u5f85\u4fee\u590d"
STATUS_FIXED = "\u5df2\u4fee\u590d"
HEADER_SCANNER = "\u626b\u63cf\u5668\u7c7b\u578b"
HEADER_SEVERITY = "\u5371\u9669\u7b49\u7ea7"
HEADER_STATUS = "\u5904\u7f6e\u72b6\u6001"


def setup_function() -> None:
    base_setup_function()


def open_first_workbook_from_zip(content: bytes):
    with zipfile.ZipFile(BytesIO(content)) as archive:
        first_name = sorted(archive.namelist())[0]
        return load_workbook(BytesIO(archive.read(first_name))), sorted(archive.namelist())


def test_export_merge_mode_creates_project_files_zip() -> None:
    client.post("/api/dev/import-assets-sample")
    client.post("/api/dev/import-vulnerabilities-sample", json={"scan_month": "2026-05"})
    response = client.get(
        "/api/exports/vulnerabilities",
        params={"scan_month": "2026-05", "handle_status": STATUS_TODO, "mode": "project-merged"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/zip"
    workbook, names = open_first_workbook_from_zip(response.content)
    assert names
    assert any(name.endswith("_2026-05_漏洞.xlsx") for name in names)
    headers = [cell.value for cell in workbook.active[1]]
    assert HEADER_SCANNER in headers
    assert HEADER_SEVERITY in headers
    assert HEADER_STATUS in headers


def test_export_split_mode_creates_project_scanner_files_zip() -> None:
    client.post("/api/dev/import-assets-sample")
    client.post("/api/dev/import-vulnerabilities-sample", json={"scan_month": "2026-05"})
    response = client.get(
        "/api/exports/vulnerabilities",
        params={"scan_month": "2026-05", "handle_status": STATUS_TODO, "mode": "project-scanner-split"},
    )
    assert response.status_code == 200
    workbook, names = open_first_workbook_from_zip(response.content)
    assert names
    assert any("_青藤云_" in name or "_阿里云_" in name or "_绿盟_" in name for name in names)
    headers = [cell.value for cell in workbook.active[1]]
    assert HEADER_SCANNER not in headers
    assert HEADER_SEVERITY in headers
    assert HEADER_STATUS in headers


def test_export_respects_status_filter() -> None:
    client.post("/api/dev/import-assets-sample")
    client.post("/api/dev/import-vulnerabilities-sample", json={"scan_month": "2026-05"})
    first_record = client.get("/api/records", params={"scan_month": "2026-05"}).json()["items"][0]
    client.patch(
        f"/api/records/{first_record['id']}",
        json={"handle_status": STATUS_FIXED, "remark": "done", "changed_by": "admin"},
    )
    response = client.get(
        "/api/exports/vulnerabilities",
        params={"scan_month": "2026-05", "handle_status": STATUS_FIXED, "mode": "project-merged"},
    )
    assert response.status_code == 200
    workbook, _ = open_first_workbook_from_zip(response.content)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert len(rows) == 2
    header_index = {name: idx for idx, name in enumerate(rows[0])}
    assert rows[1][header_index[HEADER_STATUS]] == STATUS_FIXED


def test_export_includes_severity_column_in_both_modes() -> None:
    client.post("/api/dev/import-assets-sample")
    client.post("/api/dev/import-vulnerabilities-sample", json={"scan_month": "2026-05"})
    for mode in ("project-merged", "project-scanner-split"):
        response = client.get(
            "/api/exports/vulnerabilities",
            params={"scan_month": "2026-05", "handle_status": STATUS_TODO, "mode": mode},
        )
        assert response.status_code == 200
        workbook, _ = open_first_workbook_from_zip(response.content)
        headers = [cell.value for cell in workbook.active[1]]
        assert HEADER_SEVERITY in headers
