from io import BytesIO
from pathlib import Path
import sys
import zipfile

from openpyxl import load_workbook


TESTS_DIR = Path(__file__).resolve().parent
BACKEND_DIR = TESTS_DIR.parent
ROOT_DIR = BACKEND_DIR.parent
PROTOTYPE_SRC_DIR = ROOT_DIR / "prototype" / "src"
if str(TESTS_DIR) not in sys.path:
    sys.path.insert(0, str(TESTS_DIR))
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))
if str(PROTOTYPE_SRC_DIR) not in sys.path:
    sys.path.insert(0, str(PROTOTYPE_SRC_DIR))

from app.models import Asset
from test_api import TestingSessionLocal, client, sample_path, setup_function as base_setup_function, upload_vulnerability


SCANNER_QINGTENG = "青藤云"
STATUS_FIXED = "已修复"


def setup_function() -> None:
    base_setup_function()


def workbook_from_bytes(content: bytes):
    return load_workbook(BytesIO(content))


def publish_unmatched_qingteng(scan_month: str = "2026-06") -> None:
    preview = upload_vulnerability(SCANNER_QINGTENG, scan_month, sample_path("青藤云漏洞（样例）.xlsx")).json()
    published = client.post(
        f"/api/imports/vulnerabilities/{preview['id']}/publish",
        json={"confirm_warnings": True},
    )
    assert published.status_code == 200


def test_unmatched_assets_export_supports_raw_and_dedup_modes() -> None:
    publish_unmatched_qingteng()

    raw_export = client.get(
        "/api/anomalies/unmatched-assets/export",
        params={"scan_month": "2026-06", "dedup_by_ip": "false"},
    )
    assert raw_export.status_code == 200
    raw_book = workbook_from_bytes(raw_export.content)
    raw_rows = list(raw_book.active.iter_rows(values_only=True))
    assert raw_rows[0][:4] == ("IP", "扫描器类型", "月份", "项目")
    assert len(raw_rows) >= 2

    dedup_export = client.get(
        "/api/anomalies/unmatched-assets/export",
        params={"scan_month": "2026-06", "dedup_by_ip": "true"},
    )
    assert dedup_export.status_code == 200
    dedup_book = workbook_from_bytes(dedup_export.content)
    dedup_rows = list(dedup_book.active.iter_rows(values_only=True))
    assert dedup_rows[0][:6] == ("IP", "扫描器类型", "月份", "项目", "漏洞名称", "出现次数")
    assert len(dedup_rows) >= 2


def test_unmatched_assets_rematch_updates_records_after_asset_added() -> None:
    publish_unmatched_qingteng()
    before = client.get("/api/records", params={"scan_month": "2026-06"}).json()["items"]
    target = next(item for item in before if item["asset_match_status"] == "待补充资产")
    assert target["project"] == ""

    with TestingSessionLocal() as db:
        db.add(
            Asset(
                ip=target["ip"],
                organization="新增组织",
                project="新增系统",
                workspace="新增空间",
                owner="新增负责人",
                raw_data="{}",
            )
        )
        db.commit()

    rematch = client.post(
        "/api/anomalies/unmatched-assets/rematch",
        json={"scan_month": "2026-06", "changed_by": "admin"},
    )
    assert rematch.status_code == 200
    assert rematch.json()["matched_count"] >= 1

    after = client.get("/api/records", params={"scan_month": "2026-06"}).json()["items"]
    refreshed = next(item for item in after if item["id"] == target["id"])
    assert refreshed["asset_match_status"] == "已匹配"
    assert refreshed["project"] == "新增系统"
    assert refreshed["owner"] == "新增负责人"


def test_records_can_filter_and_mark_reopened_vulnerabilities() -> None:
    client.post("/api/dev/import-assets-sample")
    client.post("/api/dev/import-vulnerabilities-sample", json={"scan_month": "2026-05"})
    may_records = client.get(
        "/api/records",
        params={"scan_month": "2026-05", "scanner_type": SCANNER_QINGTENG},
    ).json()["items"]
    target = may_records[0]
    patched = client.patch(
        f"/api/records/{target['id']}",
        json={"handle_status": STATUS_FIXED, "remark": "上月已修复", "changed_by": "admin"},
    )
    assert patched.status_code == 200

    client.post("/api/dev/import-vulnerabilities-sample", json={"scan_month": "2026-06"})
    june_records = client.get(
        "/api/records",
        params={"scan_month": "2026-06", "scanner_type": SCANNER_QINGTENG},
    ).json()["items"]
    reopened = next(item for item in june_records if item["ip"] == target["ip"] and item["vuln_name"] == target["vuln_name"])
    assert reopened["is_reopened"] is True
    assert reopened["reopened_from_status"] == STATUS_FIXED

    filtered = client.get(
        "/api/records",
        params={"scan_month": "2026-06", "scanner_type": SCANNER_QINGTENG, "reopened_only": "true"},
    )
    assert filtered.status_code == 200
    assert filtered.json()["total"] >= 1
    assert all(item["is_reopened"] is True for item in filtered.json()["items"])


def test_statistics_v11_endpoints_return_rankings_and_trends() -> None:
    client.post("/api/dev/import-assets-sample")
    client.post("/api/dev/import-vulnerabilities-sample", json={"scan_month": "2026-05"})
    client.post("/api/dev/import-vulnerabilities-sample", json={"scan_month": "2026-06"})

    top_projects = client.get("/api/statistics/top-projects", params={"scan_month_from": "2026-05", "scan_month_to": "2026-06"})
    assert top_projects.status_code == 200
    assert top_projects.json()["items"]

    top_vulns = client.get("/api/statistics/top-vulnerabilities", params={"scan_month_from": "2026-05", "scan_month_to": "2026-06"})
    assert top_vulns.status_code == 200
    assert top_vulns.json()["items"]

    monthly = client.get("/api/statistics/monthly-trend", params={"scan_month_from": "2026-05", "scan_month_to": "2026-06"})
    assert monthly.status_code == 200
    assert len(monthly.json()["items"]) == 2
    assert {"scan_month", "total", "fixed_rate", "scanner_counts"}.issubset(monthly.json()["items"][0].keys())

    reopened = client.get("/api/statistics/reopened-trend", params={"scan_month_from": "2026-05", "scan_month_to": "2026-06"})
    assert reopened.status_code == 200
    assert len(reopened.json()["items"]) == 2


def test_dashboard_drilldown_endpoints_return_record_details() -> None:
    client.post("/api/dev/import-assets-sample")
    client.post("/api/dev/import-vulnerabilities-sample", json={"scan_month": "2026-05"})

    may_records = client.get(
        "/api/records",
        params={"scan_month": "2026-05", "scanner_type": SCANNER_QINGTENG},
    ).json()["items"]
    target = may_records[0]
    client.patch(
        f"/api/records/{target['id']}",
        json={"handle_status": STATUS_FIXED, "remark": "previously fixed", "changed_by": "admin"},
    )

    client.post("/api/dev/import-vulnerabilities-sample", json={"scan_month": "2026-06"})
    june_records = client.get(
        "/api/records",
        params={"scan_month": "2026-06", "scanner_type": SCANNER_QINGTENG},
    ).json()["items"]
    june_target = next(
        item
        for item in june_records
        if item["ip"] == target["ip"] and item["port"] == target["port"] and item["vuln_name"] == target["vuln_name"]
    )

    project_detail = client.get(
        "/api/statistics/top-projects/details",
        params={"scan_month": "2026-06", "project": june_target["project"]},
    )
    assert project_detail.status_code == 200
    project_data = project_detail.json()
    assert project_data["dimension"] == "project"
    assert project_data["value"] == june_target["project"]
    assert any(item["id"] == june_target["id"] for item in project_data["items"])

    vuln_detail = client.get(
        "/api/statistics/top-vulnerabilities/details",
        params={"scan_month": "2026-06", "vuln_name": june_target["vuln_name"]},
    )
    assert vuln_detail.status_code == 200
    vuln_data = vuln_detail.json()
    assert vuln_data["dimension"] == "vuln_name"
    assert vuln_data["value"] == june_target["vuln_name"]
    assert any(item["ip"] == june_target["ip"] for item in vuln_data["items"])

    reopened_detail = client.get(
        "/api/statistics/reopened/details",
        params={"scan_month": "2026-06"},
    )
    assert reopened_detail.status_code == 200
    reopened_data = reopened_detail.json()
    assert reopened_data["dimension"] == "reopened"
    assert reopened_data["value"] == "2026-06"
    assert any(item["id"] == june_target["id"] and item["is_reopened"] is True for item in reopened_data["items"])
