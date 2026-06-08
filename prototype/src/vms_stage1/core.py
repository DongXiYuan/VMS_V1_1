from __future__ import annotations

import hashlib
import ipaddress
import json
import re
import shutil
import tempfile
import zipfile
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


STANDARD_HEADERS = [
    "序号", "扫描月份", "扫描器类型", "IP", "端口", "协议", "服务", "组织", "项目",
    "工作空间", "负责人", "危险等级", "漏洞名称", "漏洞详情", "漏洞验证",
    "漏洞修复方案", "处置状态", "备注", "上月处置状态", "上月备注", "资产匹配状态", "来源文件",
]

ANOMALY_HEADERS = ["扫描器类型", "来源文件", "IP", "异常类型", "异常详情"]


@dataclass
class StandardRecord:
    scan_month: str
    scanner_type: str
    ip: str
    port: str = ""
    protocol: str = ""
    service: str = ""
    organization: str = ""
    project: str = ""
    workspace: str = ""
    owner: str = ""
    severity: str = ""
    vuln_name: str = ""
    vuln_detail: str = ""
    verify_info: str = ""
    fix_method: str = ""
    cve: str = ""
    handle_status: str = "待修复"
    remark: str = ""
    previous_month_status: str = "无"
    previous_month_remark: str = "无"
    asset_match_status: str = "已匹配"
    source_file: str = ""

    def dedupe_key(self) -> tuple[str, str, str, str, str]:
        return (
            self.scanner_type,
            self.scan_month,
            self.ip,
            self.port,
            normalize_vuln_name(self.vuln_name),
        )

    def output_row(self) -> list[str]:
        values = asdict(self)
        return [
            values["scan_month"], values["scanner_type"], values["ip"], values["port"],
            values["protocol"], values["service"], values["organization"], values["project"],
            values["workspace"], values["owner"], values["severity"], values["vuln_name"],
            values["vuln_detail"], values["verify_info"], values["fix_method"], values["handle_status"],
            values["remark"], values["previous_month_status"], values["previous_month_remark"],
            values["asset_match_status"], values["source_file"],
        ]


@dataclass
class Anomaly:
    scanner_type: str
    source_file: str
    ip: str
    anomaly_type: str
    detail: str

    def output_row(self) -> list[str]:
        return [self.scanner_type, self.source_file, self.ip, self.anomaly_type, self.detail]


@dataclass
class ParseResult:
    scanner_type: str
    records: list[StandardRecord] = field(default_factory=list)
    anomalies: list[Anomaly] = field(default_factory=list)
    raw_rows: int = 0
    filtered_low: int = 0
    merged_duplicates: int = 0
    skipped_hosts: int = 0
    indexed_hosts: int = 0
    parsed_host_reports: int = 0


def load_config(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_port(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    try:
        return str(int(float(text)))
    except ValueError:
        return text


def normalize_vuln_name(value: Any) -> str:
    return re.sub(r"\s+", " ", normalize_text(value)).casefold()


def normalize_severity(value: Any) -> str | None:
    text = normalize_text(value)
    text = text.replace("[", "").replace("]", "").replace("【", "").replace("】", "")
    text = text.strip().casefold()
    mapping = {
        "低": "低危", "低危": "低危", "低风险": "低危", "low": "低危",
        "中": "中危", "中危": "中危", "中风险": "中危", "medium": "中危",
        "高": "高危", "高危": "高危", "高风险": "高危", "high": "高危",
        "危急": "危急", "严重": "危急", "critical": "危急",
    }
    return mapping.get(text)


def valid_ip(value: Any) -> str:
    text = normalize_text(value)
    try:
        return str(ipaddress.ip_address(text))
    except ValueError:
        return ""


def header_map(headers: Iterable[Any]) -> dict[str, int]:
    return {normalize_text(value): index for index, value in enumerate(headers) if normalize_text(value)}


def resolve_fields(headers: Iterable[Any], config: dict[str, Any]) -> dict[str, int]:
    by_name = header_map(headers)
    resolved: dict[str, int] = {}
    missing: list[str] = []
    for standard_name, original_name in config["fields"].items():
        if original_name in by_name:
            resolved[standard_name] = by_name[original_name]
        elif standard_name in config.get("required", []):
            missing.append(original_name)
    if missing:
        raise ValueError("缺少必填列: " + "、".join(missing))
    return resolved


def xlsx_rows(path: Path, config: dict[str, Any]) -> tuple[list[tuple[Any, ...]], dict[str, int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if config["sheet"] not in workbook.sheetnames:
            raise ValueError(f"缺少工作表: {config['sheet']}")
        sheet = workbook[config["sheet"]]
        rows = list(sheet.iter_rows(values_only=True))
        header_index = config["header_row"] - 1
        return rows[header_index + 1 :], resolve_fields(rows[header_index], config)
    finally:
        workbook.close()


def xls_rows(path: Path, config: dict[str, Any]) -> tuple[list[list[Any]], dict[str, int]]:
    workbook = xlrd.open_workbook(path)
    if config["sheet"] not in workbook.sheet_names():
        raise ValueError(f"缺少工作表: {config['sheet']}")
    sheet = workbook.sheet_by_name(config["sheet"])
    header_index = config["header_row"] - 1
    headers = sheet.row_values(header_index)
    rows = [sheet.row_values(index) for index in range(header_index + 1, sheet.nrows)]
    return rows, resolve_fields(headers, config)


def excel_rows(path: Path, config: dict[str, Any]) -> tuple[list[Iterable[Any]], dict[str, int]]:
    if path.suffix.lower() == ".xls":
        return xls_rows(path, config)
    return xlsx_rows(path, config)


def attach_asset(record: StandardRecord, assets: dict[str, dict[str, str]], anomalies: list[Anomaly]) -> None:
    asset = assets.get(record.ip)
    if asset:
        record.organization = asset["organization"]
        record.project = asset["project"]
        record.workspace = asset["workspace"]
        record.owner = asset["owner"]
        return
    record.asset_match_status = "待补充资产"
    anomalies.append(Anomaly(record.scanner_type, record.source_file, record.ip, "资产未匹配", "IP 不在资产表中"))


def dedupe(records: list[StandardRecord]) -> tuple[list[StandardRecord], int]:
    unique: dict[tuple[str, str, str, str, str], StandardRecord] = {}
    for record in records:
        unique[record.dedupe_key()] = record
    return list(unique.values()), len(records) - len(unique)


def history_key(record: StandardRecord) -> tuple[str, str, str, str]:
    return (
        record.scanner_type,
        record.ip,
        record.port,
        normalize_vuln_name(record.vuln_name),
    )


def attach_previous_month(records: list[StandardRecord], previous_records: Iterable[StandardRecord]) -> None:
    previous_by_key = {history_key(record): record for record in previous_records}
    for record in records:
        previous = previous_by_key.get(history_key(record))
        if previous:
            record.previous_month_status = previous.handle_status or "无"
            record.previous_month_remark = previous.remark or "无"


def parse_assets(path: Path, config: dict[str, Any]) -> dict[str, dict[str, str]]:
    rows, fields = excel_rows(path, config)
    assets: dict[str, dict[str, str]] = {}
    for row in rows:
        ip = valid_ip(row[fields["ip"]])
        if not ip:
            continue
        assets[ip] = {
            "organization": normalize_text(row[fields["organization"]]),
            "project": normalize_text(row[fields["project"]]),
            "workspace": normalize_text(row[fields["workspace"]]),
            "owner": normalize_text(row[fields["owner"]]),
        }
    return assets


def parse_simple_xlsx(
    path: Path, scan_month: str, scanner_type: str, config: dict[str, Any], assets: dict[str, dict[str, str]]
) -> ParseResult:
    result = ParseResult(scanner_type)
    try:
        rows, fields = excel_rows(path, config)
    except ValueError as error:
        result.anomalies.append(Anomaly(scanner_type, path.name, "", "文件结构异常", str(error)))
        return result
    for row in rows:
        if not normalize_text(row[fields["vuln_name"]]):
            continue
        result.raw_rows += 1
        ip = valid_ip(row[fields["ip"]])
        if not ip:
            result.anomalies.append(Anomaly(scanner_type, path.name, normalize_text(row[fields["ip"]]), "IP 异常", "IP 为空或格式错误"))
            continue
        severity = normalize_severity(row[fields["severity"]])
        if not severity:
            result.anomalies.append(Anomaly(scanner_type, path.name, ip, "危险等级异常", normalize_text(row[fields["severity"]])))
            continue
        if severity == "低危":
            result.filtered_low += 1
            continue
        record = StandardRecord(
            scan_month=scan_month,
            scanner_type=scanner_type,
            ip=ip,
            severity=severity,
            vuln_name=normalize_text(row[fields["vuln_name"]]),
            vuln_detail=normalize_text(row[fields["vuln_detail"]]) if "vuln_detail" in fields else "",
            verify_info=normalize_text(row[fields["verify_info"]]) if "verify_info" in fields else "",
            fix_method=normalize_text(row[fields["fix_method"]]) if "fix_method" in fields else "",
            cve=normalize_text(row[fields["cve"]]) if "cve" in fields else "",
            source_file=path.name,
        )
        attach_asset(record, assets, result.anomalies)
        result.records.append(record)
    result.records, result.merged_duplicates = dedupe(result.records)
    return result


def safe_extract(zip_path: Path, target: Path) -> None:
    with zipfile.ZipFile(zip_path) as archive:
        for member in archive.infolist():
            destination = (target / member.filename).resolve()
            if target.resolve() not in destination.parents and destination != target.resolve():
                raise ValueError(f"ZIP 路径越界: {member.filename}")
        archive.extractall(target)


def find_host_report(directory: Path, ip: str) -> Path | None:
    exact = directory / f"{ip}.xls"
    if exact.exists():
        return exact
    candidates = sorted(directory.glob(f"{ip}*.xls"))
    return candidates[0] if candidates else None


def parse_nsfocus_host(
    path: Path, scan_month: str, config: dict[str, Any], assets: dict[str, dict[str, str]]
) -> ParseResult:
    result = ParseResult("绿盟")
    ip = valid_ip(path.name.split("（", 1)[0].replace(".xls", ""))
    try:
        rows, fields = xls_rows(path, config)
    except ValueError as error:
        result.anomalies.append(Anomaly("绿盟", path.name, ip, "主机报告结构异常", str(error)))
        return result
    current_port = current_protocol = current_service = ""
    for row in rows:
        vuln_name = normalize_text(row[fields["vuln_name"]])
        port_value = normalize_port(row[fields["port"]]) if "port" in fields else ""
        protocol_value = normalize_text(row[fields["protocol"]]) if "protocol" in fields else ""
        service_value = normalize_text(row[fields["service"]]) if "service" in fields else ""
        if not any((vuln_name, port_value, protocol_value, service_value)):
            current_port = current_protocol = current_service = ""
            continue
        if port_value:
            current_port = port_value
        if protocol_value:
            current_protocol = protocol_value
        if service_value:
            current_service = service_value
        if not vuln_name:
            continue
        result.raw_rows += 1
        severity = normalize_severity(row[fields["severity"]])
        if not severity:
            result.anomalies.append(Anomaly("绿盟", path.name, ip, "危险等级异常", normalize_text(row[fields["severity"]])))
            continue
        if severity == "低危":
            result.filtered_low += 1
            continue
        record = StandardRecord(
            scan_month=scan_month,
            scanner_type="绿盟",
            ip=ip,
            port=current_port,
            protocol=current_protocol,
            service=current_service,
            severity=severity,
            vuln_name=vuln_name,
            vuln_detail=normalize_text(row[fields["vuln_detail"]]) if "vuln_detail" in fields else "",
            verify_info=normalize_text(row[fields["verify_info"]]) if "verify_info" in fields else "",
            fix_method=normalize_text(row[fields["fix_method"]]) if "fix_method" in fields else "",
            cve=normalize_text(row[fields["cve"]]) if "cve" in fields else "",
            source_file=path.name,
        )
        attach_asset(record, assets, result.anomalies)
        result.records.append(record)
    result.records, result.merged_duplicates = dedupe(result.records)
    return result


def parse_nsfocus(
    source: Path, scan_month: str, index_config: dict[str, Any], host_config: dict[str, Any], assets: dict[str, dict[str, str]]
) -> ParseResult:
    result = ParseResult("绿盟")
    temporary: tempfile.TemporaryDirectory[str] | None = None
    if source.is_file() and source.suffix.lower() == ".zip":
        temporary = tempfile.TemporaryDirectory(prefix="vms_nsfocus_")
        directory = Path(temporary.name)
        safe_extract(source, directory)
    else:
        directory = source
    indexes = sorted(directory.glob("index*.xls"))
    if not indexes:
        result.anomalies.append(Anomaly("绿盟", source.name, "", "缺少索引", "未找到 index.xls"))
        return result
    index_path = indexes[0]
    try:
        rows, fields = xls_rows(index_path, index_config)
    except ValueError as error:
        result.anomalies.append(Anomaly("绿盟", index_path.name, "", "索引结构异常", str(error)))
        return result
    for row in rows:
        ip = valid_ip(row[fields["ip"]])
        if not ip:
            continue
        result.indexed_hosts += 1
        high = float(row[fields["high_count"]] or 0)
        medium = float(row[fields["medium_count"]] or 0)
        if high + medium <= 0:
            result.skipped_hosts += 1
            continue
        report = find_host_report(directory, ip)
        if not report:
            result.anomalies.append(Anomaly("绿盟", index_path.name, ip, "缺少主机报告", f"索引存在中高危漏洞，但未找到 {ip}.xls"))
            continue
        host_result = parse_nsfocus_host(report, scan_month, host_config, assets)
        result.records.extend(host_result.records)
        result.anomalies.extend(host_result.anomalies)
        result.raw_rows += host_result.raw_rows
        result.filtered_low += host_result.filtered_low
        result.merged_duplicates += host_result.merged_duplicates
        result.parsed_host_reports += 1
    result.records, merged = dedupe(result.records)
    result.merged_duplicates += merged
    if temporary:
        temporary.cleanup()
    return result


def write_workbook(output_path: Path, records: list[StandardRecord], anomalies: list[Anomaly], results: list[ParseResult]) -> None:
    workbook = Workbook()
    standard = workbook.active
    standard.title = "标准漏洞清单"
    standard.append(STANDARD_HEADERS)
    for index, record in enumerate(records, 1):
        standard.append([index, *record.output_row()])

    anomaly_sheet = workbook.create_sheet("异常数据")
    anomaly_sheet.append(ANOMALY_HEADERS)
    for anomaly in anomalies:
        anomaly_sheet.append(anomaly.output_row())

    stats = workbook.create_sheet("导入统计")
    stats.append(["扫描器类型", "原始有效行", "标准记录数", "低危过滤数", "重复合并数", "异常数", "索引主机数", "解析主机报告数", "跳过主机数"])
    for result in results:
        stats.append([
            result.scanner_type, result.raw_rows, len(result.records), result.filtered_low,
            result.merged_duplicates, len(result.anomalies), result.indexed_hosts,
            result.parsed_host_reports, result.skipped_hosts,
        ])
    stats.append(["合计", sum(r.raw_rows for r in results), len(records), sum(r.filtered_low for r in results),
                  sum(r.merged_duplicates for r in results), len(anomalies), "", "", ""])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        for column_cells in sheet.columns:
            max_length = max(len(normalize_text(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 45)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_json_summary(output_path: Path, records: list[StandardRecord], anomalies: list[Anomaly], results: list[ParseResult]) -> None:
    summary = {
        "standard_record_count": len(records),
        "anomaly_count": len(anomalies),
        "scanners": {
            result.scanner_type: {
                "raw_rows": result.raw_rows,
                "standard_records": len(result.records),
                "filtered_low": result.filtered_low,
                "merged_duplicates": result.merged_duplicates,
                "anomalies": len(result.anomalies),
                "indexed_hosts": result.indexed_hosts,
                "parsed_host_reports": result.parsed_host_reports,
                "skipped_hosts": result.skipped_hosts,
            }
            for result in results
        },
    }
    output_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def first_existing_path(*candidates: Path) -> Path:
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError(f"未找到样例文件: {[str(candidate) for candidate in candidates]}")


def resolve_nsfocus_source(samples_dir: Path) -> Path:
    return first_existing_path(
        samples_dir / "绿盟",
        samples_dir / "绿盟.zip",
        samples_dir,
    )


def run_pipeline(
    samples_dir: Path,
    output_dir: Path,
    scan_month: str,
    config_path: Path,
    previous_records: Iterable[StandardRecord] = (),
) -> dict[str, Any]:
    config = load_config(config_path)
    assets = parse_assets(first_existing_path(samples_dir / "资产表（样例）.xlsx"), config["assets"])
    results = [
        parse_simple_xlsx(first_existing_path(samples_dir / "青藤云漏洞（样例）.xlsx"), scan_month, "青藤云", config["qingteng"], assets),
        parse_simple_xlsx(first_existing_path(samples_dir / "阿里云漏洞（样例）.xlsx"), scan_month, "阿里云", config["aliyun"], assets),
        parse_nsfocus(resolve_nsfocus_source(samples_dir), scan_month, config["nsfocus_index"], config["nsfocus_host"], assets),
    ]
    records = [record for result in results for record in result.records]
    attach_previous_month(records, previous_records)
    anomalies = [anomaly for result in results for anomaly in result.anomalies]
    output_dir.mkdir(parents=True, exist_ok=True)
    write_workbook(output_dir / "阶段1_标准漏洞清单.xlsx", records, anomalies, results)
    write_json_summary(output_dir / "阶段1_导入统计.json", records, anomalies, results)
    return {"assets": assets, "records": records, "anomalies": anomalies, "results": results}
